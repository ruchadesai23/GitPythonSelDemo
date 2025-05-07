from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
import openpyxl

import os
import time


def wait_for_file_download(filepath, timeout=20):
    """Wait until the file is downloaded and accessible."""
    for _ in range(timeout):
        if os.path.exists(filepath):
            try:
                open(filepath).close()
                return True
            except:
                pass
        time.sleep(1)
    raise FileNotFoundError(f"File not downloaded: {filepath}")


def update_data_in_excel(filepath, searchterm, column_name, new_value):  # these are the required parameters
    book = openpyxl.load_workbook("C:/Users/rucha/Downloads/download.xlsx")
    sheet = book.active
    Dict = {}

    # getting the value of column
    for i in range(1, sheet.max_row + 1):  # this will iterate through each row
        if sheet.cell(row=1, column=i).value == column_name:  # this will get the column name from the method
            # above we are retrieving value of the column till we find price.
            Dict["col"] = i
            # above we are adding the value of price column to dictionary. this will be used in the testcase

    # getting the value of row
    for i in range(1, sheet.max_row + 1):  # this will iterate through each row
        for j in range(1, sheet.max_column + 1):  # this will iterate through each column
            if sheet.cell(row=i, column=j).value == searchterm:  # # this will get the search term from the method
                Dict["row"] = i  # store the row value in the dictionary

    sheet.cell(row=Dict["row"], column=Dict["col"]).value = new_value  # this will get the new value from the method
    book.save(filepath)  # saving the changes in the excel #filepath is taken from the method


file_path = "C:/Users/rucha/Downloads/download.xlsx"
fruit_name = "Apple"
newValue = "999"

driver = webdriver.Chrome()
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/upload-download-test/")

# downloaded the file
driver.find_element(By.ID, "downloadButton").click()

# wait for the download to complete
wait_for_file_download(file_path)

# update excel data
update_data_in_excel(file_path, fruit_name, "price", newValue)  # this is the method with values


# uploaded the file
file_input = driver.find_element(By.XPATH, "//input[@type='file']")
file_input.send_keys(file_path)  # used sendkeys since it will be a window screen

wait = WebDriverWait(driver, 5)
toast_locator = (By.XPATH, "//div[@class='Toastify__toast-body']/div[2]")
wait.until(EC.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)
price_column = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
# below is a smart xpath where fruit and price column is selected dynamically
actual_price = driver.find_element(By.XPATH,"//div[text()='" + fruit_name + "']/parent::div/parent::div/div[@id='cell-" + price_column + "-undefined']").text
print(actual_price)
time.sleep(5)
assert actual_price == newValue
