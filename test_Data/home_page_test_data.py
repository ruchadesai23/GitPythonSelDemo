import openpyxl


class HomePageTestData:
    test_home_page_data = [{"first_name": "Rucha", "email": "hello@gmail.com", "password": "12345", "gender": "Female"},
                            {"first_name": "siddhesh", "email": "us@yopmail.com", "password": "654123","gender": "Male"}]

    # here we have created a class for test data for home page test cases.
    # we have added all the data in dictionary format and we will import this class in the test case.

    @staticmethod  # in order to avoid creating an object in the testcase we are defining it as static method
    def get_exceltest_data(testcase_name):  # here we removed self as it is only required for nonstatic methods
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\rucha\\Documents\\PythonDemo.xlsx")
        sheet = book.active

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == testcase_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict] #returning dictionary as a list, since the data has to be a list
