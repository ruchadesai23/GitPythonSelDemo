import openpyxl

# first go to the book
book = openpyxl.load_workbook("C:\\Users\\rucha\\Documents\\PythonDemo.xlsx")
# go to the active sheet
sheet = book.active
# cell = sheet.cell(row=1, column=2)  # to read from the excelsheet
# print(cell.value)  # this will get the value of the cell
# print(sheet["A5"].value)  # this acts same as above only we are mentioning the cell name from the excel
#
# sheet.cell(row=2, column=2).value = "Rahul"  # this will write back to the excel sheet
# print(sheet.cell(row=2, column=2).value)
#
# print(sheet.max_row)  # get the maximum no of rows
# print(sheet.max_column)  # get the maximum no of columns

# to print all the rows in the excel
Dict = {}
for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "Testcase 2":
        for j in range(2, sheet.max_column + 1):
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value



# dict["Firstname"] ="Rucha"

# for i in range(1, sheet.max_row + 1):  # this will get values from all the rows
#     # above we are mentioning +1 to consider the last row as the range will only print till the second last row
#     if sheet.cell(row=i, column=1).value == "Testcase 1": # this is for selecting data for specific testcase
#         for j in range(2, sheet.max_column + 1):
#         # above 2 is the index to skip the tescase name and just take values from columns
#         # for j in range(1, sheet.max_column + 1): # here 1 is the 1st cell of the column.
#             print(sheet.cell(row=i, column=j).value)  # gets the value in each cell in all rows in column 1